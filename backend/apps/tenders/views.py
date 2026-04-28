"""
招标项目模块 - 视图（优化版）
"""
import logging
from rest_framework import status, generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.db.models import Q

from .models import TenderSource, TenderProject, TenderFile, TenderKeyword, CrawlerTask

logger = logging.getLogger(__name__)

from .serializers import (
    TenderSourceSerializer, TenderProjectListSerializer, TenderProjectDetailSerializer,
    TenderProjectCreateSerializer, TenderProjectUpdateSerializer, TenderFileSerializer,
    TenderKeywordSerializer, CrawlerTaskSerializer,
    TenderBatchDeleteSerializer, TenderBatchUpdateSerializer
)
from .services import TenderService, TenderKeywordService, CrawlerTaskService
from utils.permissions import IsAdminUser
from utils.responses import UnifiedResponse
from core.cache import cache_result
from core.pagination import StandardPagination


class TenderSourceListView(generics.ListCreateAPIView):
    """
    招标来源列表视图
    """
    queryset = TenderSource.objects.all()
    serializer_class = TenderSourceSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAuthenticated(), IsAdminUser()]
        return [IsAuthenticated()]

class TenderProjectListView(generics.ListCreateAPIView):
    """
    招标项目列表视图
    """
    permission_classes = [IsAuthenticated]
    serializer_class = TenderProjectListSerializer
    pagination_class = StandardPagination

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return TenderProjectCreateSerializer
        return TenderProjectListSerializer

    def get_queryset(self):
        """
        支持多条件搜索
        """
        return TenderService.search_tenders(
            keyword=self.request.query_params.get('keyword'),
            region=self.request.query_params.get('region'),
            industry=self.request.query_params.get('industry'),
            status=self.request.query_params.get('status'),
            start_date=self.request.query_params.get('start_date'),
            end_date=self.request.query_params.get('end_date'),
            is_favorite=self.request.query_params.get('is_favorite'),
            is_read=self.request.query_params.get('is_read'),
            source_name=self.request.query_params.get('source_name')
        )

    def perform_create(self, serializer):
        tender = serializer.save(created_by=self.request.user)


class TenderProjectDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    招标项目详情视图
    """
    queryset = TenderProject.objects.select_related('source').prefetch_related('files')
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return TenderProjectUpdateSerializer
        return TenderProjectDetailSerializer

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return UnifiedResponse.success(data=serializer.data)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return UnifiedResponse.success(data=serializer.data, message='更新成功')


class TenderProjectBatchView(APIView):
    """
    招标项目批量操作视图
    """
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def delete(self, request):
        """
        批量删除
        """
        serializer = TenderBatchDeleteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        tender_ids = serializer.validated_data.get('ids')
        count = TenderService.batch_delete(tender_ids, request.user)
        
        return UnifiedResponse.success(
            data={'deleted_count': count},
            message=f'成功删除 {count} 条记录'
        )

    @transaction.atomic
    def patch(self, request):
        """
        批量更新状态
        """
        serializer = TenderBatchUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        tender_ids = serializer.validated_data.get('ids')
        new_status = serializer.validated_data.get('status')
        
        count = TenderService.batch_update_status(tender_ids, new_status, request.user)
        
        return UnifiedResponse.success(
            data={'updated_count': count},
            message=f'成功更新 {count} 条记录'
        )


class TenderProjectFavoriteView(APIView):
    """
    招标项目收藏视图
    """
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        """
        收藏/取消收藏
        """
        try:
            is_favorite = TenderService.toggle_favorite(pk)
            return UnifiedResponse.success(
                data={'is_favorite': is_favorite},
                message='收藏成功' if is_favorite else '取消收藏成功'
            )
        except TenderProject.DoesNotExist:
            return UnifiedResponse.error(message='招标项目不存在', status_code=status.HTTP_404_NOT_FOUND)

class TenderKeywordListView(generics.ListCreateAPIView):
    """
    招标关键词列表视图
    """
    serializer_class = TenderKeywordSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardPagination

    def get_queryset(self):
        category = self.request.query_params.get('category')
        return TenderKeywordService.get_active_keywords(category)

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class TenderKeywordDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    招标关键词详情视图
    """
    queryset = TenderKeyword.objects.all()
    serializer_class = TenderKeywordSerializer
    permission_classes = [IsAuthenticated]

class TenderStatisticsView(APIView):
    """
    招标统计视图
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        获取招标统计数据
        """
        data = TenderService.get_statistics()
        return UnifiedResponse.success(data=data)


class TenderTrendView(APIView):
    """
    招标趋势视图
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.utils import timezone
        from datetime import timedelta
        from django.db.models import Count
        from django.db.models.functions import TruncDate

        query_params = getattr(request, 'query_params', None) or request.GET
        days = int(query_params.get('days', 7))
        days = min(max(days, 1), 90)

        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days - 1)

        trend_data = (
            TenderProject.objects.filter(
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            )
            .annotate(date=TruncDate('created_at'))
            .values('date')
            .annotate(count=Count('id'))
            .order_by('date')
        )

        date_map = {item['date']: item['count'] for item in trend_data}

        labels = []
        data = []
        for i in range(days - 1, -1, -1):
            d = end_date - timedelta(days=i)
            labels.append(f'{d.month}/{d.day}')
            data.append(date_map.get(d, 0))

        return UnifiedResponse.success(data={'labels': labels, 'data': data})

class TenderSourceContentView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        try:
            tender = TenderProject.objects.get(pk=pk)
        except TenderProject.DoesNotExist:
            return UnifiedResponse.error(message='招标项目不存在', status_code=status.HTTP_404_NOT_FOUND)

        if not tender.source_url:
            return UnifiedResponse.error(message='该招标项目没有来源链接')

        if tender.description:
            return UnifiedResponse.success(data={
                'title': tender.title,
                'content': tender.description,
                'source_url': tender.source_url,
                'from_cache': True
            })

        content = self._fetch_content(tender.source_url)
        if content:
            TenderProject.objects.filter(pk=pk).update(description=content)
            return UnifiedResponse.success(data={
                'title': tender.title,
                'content': content,
                'source_url': tender.source_url,
                'from_cache': False
            })

        return UnifiedResponse.error(message='无法获取公告内容，请尝试直接访问原文链接')

    def _fetch_content(self, url):
        content = self._fetch_with_selenium(url)
        if content:
            return content

        content = self._fetch_with_requests(url)
        if content:
            return content

        return None

    def _fetch_with_selenium(self, url):
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.chrome.service import Service
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            import time

            options = Options()
            options.add_argument('--headless=new')
            options.add_argument('--disable-gpu')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_argument('--window-size=1920,1080')
            options.add_experimental_option('excludeSwitches', ['enable-automation'])

            try:
                service = Service()
                driver = webdriver.Chrome(service=service, options=options)
            except Exception:
                try:
                    import shutil
                    chrome_path = shutil.which('chromedriver')
                    if chrome_path:
                        service = Service(executable_path=chrome_path)
                        driver = webdriver.Chrome(service=service, options=options)
                    else:
                        return None
                except Exception:
                    return None

            try:
                driver.get(url)
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.TAG_NAME, 'body'))
                )
                time.sleep(3)

                content_html = None

                try:
                    iframes = driver.find_elements(By.TAG_NAME, 'iframe')
                    for iframe in iframes:
                        iframe_class = iframe.get_attribute('class') or ''
                        iframe_src = iframe.get_attribute('src') or ''
                        if 'content' in iframe_class or 'detail' in iframe_class or 'mapFrame' in iframe_class:
                            driver.switch_to.frame(iframe)
                            try:
                                iframe_body = driver.find_element(By.TAG_NAME, 'body')
                                iframe_html = iframe_body.get_attribute('innerHTML')
                                if iframe_html and len(iframe_html.strip()) > 100:
                                    content_html = iframe_html
                                    break
                            finally:
                                driver.switch_to.default_content()
                except Exception:
                    driver.switch_to.default_content()

                if not content_html:
                    try:
                        iframes = driver.find_elements(By.TAG_NAME, 'iframe')
                        for iframe in iframes:
                            driver.switch_to.frame(iframe)
                            try:
                                iframe_body = driver.find_element(By.TAG_NAME, 'body')
                                iframe_html = iframe_body.get_attribute('innerHTML')
                                if iframe_html and len(iframe_html.strip()) > 100:
                                    content_html = iframe_html
                                    break
                            finally:
                                driver.switch_to.default_content()
                    except Exception:
                        driver.switch_to.default_content()

                if not content_html:
                    content_selectors = [
                        '.detail-content', '.article-content', '.content',
                        '.main-content', '#detailContent', '.notice-content',
                        '.el-main', '.page-content', '.announcement-content',
                        'article', '.detail', '.news-content',
                    ]
                    for selector in content_selectors:
                        try:
                            elements = driver.find_elements(By.CSS_SELECTOR, selector)
                            for elem in elements:
                                html = elem.get_attribute('innerHTML')
                                if html and len(html.strip()) > 100:
                                    content_html = html
                                    break
                            if content_html:
                                break
                        except Exception:
                            continue

                if not content_html:
                    body = driver.find_element(By.TAG_NAME, 'body')
                    content_html = body.get_attribute('innerHTML')

                return content_html

            finally:
                driver.quit()

        except Exception as e:
            logger.error(f"Selenium获取公告内容失败: {e}")
            return None

    def _fetch_with_requests(self, url):
        try:
            import requests as req
            from bs4 import BeautifulSoup
            import urllib3
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            }
            resp = req.get(url, headers=headers, timeout=15, verify=False)

            if resp.status_code != 200:
                return None

            soup = BeautifulSoup(resp.text, 'html.parser')

            for tag in soup(['script', 'style', 'nav', 'header', 'footer', 'iframe']):
                tag.decompose()

            content_selectors = [
                '.detail-content', '.article-content', '.content',
                '.main-content', '#detailContent', '.notice-content',
                'article', '.detail', '.news-content',
            ]

            for selector in content_selectors:
                elem = soup.select_one(selector)
                if elem:
                    html = str(elem)
                    if len(html.strip()) > 100:
                        return html

            body = soup.find('body')
            if body:
                return str(body)

            return None

        except Exception as e:
            logger.error(f"Requests获取公告内容失败: {e}")
            return None


class CrawlSyncView(APIView):
    """
    采集数据同步视图
    将 crawler 采集的数据同步到 tender_projects
    """
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        """
        获取同步状态
        """
        from .services import CrawlToTenderSyncService
        status = CrawlToTenderSyncService.get_sync_status()
        return UnifiedResponse.success(data=status)

    def post(self, request):
        """
        执行同步
        """
        from .services import CrawlToTenderSyncService

        limit = request.data.get('limit')
        if limit:
            limit = int(limit)

        try:
            result = CrawlToTenderSyncService.sync_all(limit=limit)
            return UnifiedResponse.success(
                data=result,
                message=f"同步完成: 新增 {result['created']} 条, 更新 {result['updated']} 条, 跳过 {result['skipped']} 条"
            )
        except Exception as e:
            logger.error(f"同步失败: {str(e)}")
            return UnifiedResponse.error(message='同步失败，请稍后重试')


class CrawlDataStatisticsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.utils import timezone
        from datetime import timedelta
        from django.db.models import Count, Case, When, IntegerField
        from django.db.models.functions import TruncDate
        from apps.crawler.models import CrawlResult, CrawlSession, CrawlSchedule, CrawlScheduleLog, WebsiteTemplate
        from apps.crawler.scheduler_models import CrawlSchedule as ScheduleModel
        from .services import CrawlToTenderSyncService

        query_params = getattr(request, 'query_params', None) or request.GET
        days = int(query_params.get('days', 7))
        days = min(max(days, 1), 90)

        today = timezone.now().date()
        yesterday = today - timedelta(days=1)

        templates = WebsiteTemplate.objects.all()
        source_stats = []
        anomalies = []

        for template in templates:
            tender_source = TenderSource.objects.filter(code=template.code).first()

            total_count = TenderProject.objects.filter(source=tender_source).count() if tender_source else 0
            today_count = TenderProject.objects.filter(
                source=tender_source,
                created_at__date=today
            ).count() if tender_source else 0
            yesterday_count = TenderProject.objects.filter(
                source=tender_source,
                created_at__date=yesterday
            ).count() if tender_source else 0

            crawl_results = CrawlResult.objects.filter(session__website_template=template)
            matched_count = crawl_results.filter(status__in=['matched', 'synced']).count()
            ignored_count = crawl_results.filter(status='ignored').count()
            pending_count = crawl_results.filter(status__in=['pending', 'processed']).count()

            deleted_count = TenderProject.objects.filter(
                source=tender_source,
                is_deleted=True
            ).count() if tender_source else 0

            last_session = CrawlSession.objects.filter(
                website_template=template
            ).order_by('-created_at').first()
            last_crawl_at = last_session.created_at.strftime('%Y-%m-%d %H:%M:%S') if last_session else None

            source_type = tender_source.source_type if tender_source else 'other'

            source_stats.append({
                'id': template.id,
                'name': template.name,
                'code': template.code,
                'source_type': source_type,
                'today_count': today_count,
                'yesterday_count': yesterday_count,
                'total_count': total_count,
                'matched_count': matched_count,
                'ignored_count': ignored_count,
                'pending_count': pending_count,
                'deleted_count': deleted_count,
                'last_crawl_at': last_crawl_at,
                'is_active': template.is_active,
            })

            if yesterday_count > 0:
                change_rate = abs(today_count - yesterday_count) / yesterday_count
                if change_rate > 0.5:
                    direction = '上升' if today_count > yesterday_count else '下降'
                    anomalies.append({
                        'source_name': template.name,
                        'type': 'data_fluctuation',
                        'message': f'采集量{direction}幅度较大（{direction}{change_rate:.0%}），昨日{yesterday_count}条，今日{today_count}条',
                        'yesterday_count': yesterday_count,
                        'today_count': today_count,
                        'change_rate': round(change_rate, 2),
                    })

            if today_count == 0 and template.is_active:
                recent_sessions = CrawlSession.objects.filter(
                    website_template=template,
                    status__in=['failed', 'completed'],
                    created_at__date__gte=today - timedelta(days=3)
                )
                failed_count = recent_sessions.filter(status='failed').count()
                if failed_count > 0:
                    anomalies.append({
                        'source_name': template.name,
                        'type': 'crawl_error',
                        'message': f'近3天有{failed_count}次采集失败，今日无新数据',
                        'failed_count': failed_count,
                    })

            if pending_count > 100:
                anomalies.append({
                    'source_name': template.name,
                    'type': 'pending_backlog',
                    'message': f'待处理数据积压{pending_count}条',
                    'pending_count': pending_count,
                })

        total_crawled = TenderProject.objects.filter(source__isnull=False).count()
        today_crawled = TenderProject.objects.filter(
            source__isnull=False,
            created_at__date=today
        ).count()
        total_unmatched = CrawlResult.objects.filter(status='ignored').count()
        total_deleted = TenderProject.objects.filter(is_deleted=True).count()

        sync_status = CrawlToTenderSyncService.get_sync_status()
        sync_pending = max(0, sync_status.get('crawl_pending', 0))

        end_date = today
        start_date = end_date - timedelta(days=days - 1)

        trend_data = (
            TenderProject.objects.filter(
                source__isnull=False,
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            )
            .annotate(date=TruncDate('created_at'))
            .values('date')
            .annotate(count=Count('id'))
            .order_by('date')
        )

        date_map = {item['date']: item['count'] for item in trend_data}
        labels = []
        trend_counts = []
        for i in range(days - 1, -1, -1):
            d = end_date - timedelta(days=i)
            labels.append(f'{d.month}/{d.day}')
            trend_counts.append(date_map.get(d, 0))

        overview = {
            'total_sources': templates.count(),
            'active_sources': templates.filter(is_active=True).count(),
            'total_crawled': total_crawled,
            'today_crawled': today_crawled,
            'total_unmatched': total_unmatched,
            'total_deleted': total_deleted,
            'sync_pending': sync_pending,
            'sync_synced': sync_status.get('crawl_synced', 0),
        }

        data = {
            'overview': overview,
            'source_stats': source_stats,
            'trend': {
                'labels': labels,
                'data': trend_counts,
            },
            'anomalies': anomalies,
        }

        return UnifiedResponse.success(data=data)


class CrawlDataExportView(APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def get(self, request):
        import io
        from django.http import HttpResponse
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError as e:
            return UnifiedResponse.error(message=f'请先安装openpyxl: pip install openpyxl ({str(e)})')

        from django.utils import timezone
        from datetime import timedelta
        from apps.crawler.models import CrawlResult, CrawlSession, WebsiteTemplate

        today = timezone.now().date()
        templates = WebsiteTemplate.objects.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '采集数据统计'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = ['网站名称', '网站编码', '类型', '今日采集', '昨日采集', '累计采集', '已匹配', '已忽略', '待处理', '已删除', '最后采集时间']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, template in enumerate(templates, 2):
            tender_source = TenderSource.objects.filter(code=template.code).first()
            today_count = TenderProject.objects.filter(source=tender_source, created_at__date=today).count() if tender_source else 0
            yesterday_count = TenderProject.objects.filter(source=tender_source, created_at__date=today - timedelta(days=1)).count() if tender_source else 0
            total_count = TenderProject.objects.filter(source=tender_source).count() if tender_source else 0
            crawl_results = CrawlResult.objects.filter(session__website_template=template)
            matched_count = crawl_results.filter(status__in=['matched', 'synced']).count()
            ignored_count = crawl_results.filter(status='ignored').count()
            pending_count = crawl_results.filter(status__in=['pending', 'processed']).count()
            deleted_count = TenderProject.objects.filter(source=tender_source, is_deleted=True).count() if tender_source else 0

            last_session = CrawlSession.objects.filter(
                website_template=template
            ).order_by('-created_at').first()
            last_crawl_at = last_session.created_at.strftime('%Y-%m-%d %H:%M:%S') if last_session else '-'

            source_type = tender_source.source_type if tender_source else 'other'

            row_data = [
                template.name, template.code, source_type,
                today_count, yesterday_count, total_count,
                matched_count, ignored_count, pending_count,
                deleted_count, last_crawl_at
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 4, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'采集数据统计_{today.strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
